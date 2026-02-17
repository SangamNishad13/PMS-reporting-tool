from pathlib import Path
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
MODULES_DIR = ROOT / "modules"
OUTPUT = ROOT / "PMS_Manual_Test_Scenarios_Module_Page_Flow.xlsx"


def discover_pages():
    pages = []
    for b, _, fs in os.walk(MODULES_DIR):
        for f in fs:
            if f.endswith(".php"):
                pages.append((Path(b) / f).relative_to(ROOT).as_posix())
    return sorted(pages)


def group_modules(pages):
    m = {}
    for p in pages:
        parts = p.split("/")
        mod = parts[1] if len(parts) > 2 else "core"
        m.setdefault(mod, []).append(p)
    return dict(sorted(m.items()))


def get_flows():
    return [
        ("F01", "Authentication & Access", "Secure login/logout and guards", "auth,core", "High"),
        ("F02", "User & Session Management", "Manage users/sessions/activity", "admin", "High"),
        ("F03", "Project Lifecycle", "Create/edit/archive/duplicate/delete", "projects,project_lead,admin", "High"),
        ("F04", "Assignment & Page Management", "Team/page/env assignments", "projects,project_lead,qa", "High"),
        ("F05", "Testing Execution", "AT/FT/QA task workflow", "at_tester,ft_tester,qa", "High"),
        ("F06", "Issue Management", "Issue lifecycle and exports", "projects,admin", "High"),
        ("F07", "Status & Hours", "Daily status and production logs", "core,admin", "High"),
        ("F08", "Chat & Collaboration", "Chat/messages/mentions/files", "chat,core", "Medium"),
        ("F09", "Reports & Exports", "Reporting and export correctness", "reports,admin", "Medium"),
        ("F10", "Admin Masters", "Client/env/status master data", "admin", "Medium"),
        ("F11", "Uploads & Assets", "Asset/upload security + cleanup", "projects,admin,chat", "High"),
        ("F12", "Profile & Notifications", "Assigned projects/tasks visibility", "core,admin", "High"),
        ("F13", "Generic Tasks", "Task/category management", "generic_tasks", "Low"),
        ("F14", "Security & Regression", "RBAC, confirmations, regressions", "all", "High"),
    ]


def get_cases():
    rows = []
    add = rows.append
    add(("F01", "auth", "modules/auth/login.php", "Valid login", "Active user", "Login with valid credentials", "Redirect to dashboard", "High", "Positive", "All"))
    add(("F01", "auth", "modules/auth/login.php", "Invalid login", "Existing user", "Login with wrong password", "Error shown, no session", "High", "Negative", "All"))
    add(("F01", "auth", "modules/auth/logout.php", "Logout", "Logged in", "Click logout", "Session ended and redirect to login", "High", "Positive", "All"))
    add(("F01", "core", "Protected URLs", "Unauthorized URL access", "Not logged in", "Open protected URL directly", "Redirect to login", "High", "Security", "Guest"))

    add(("F02", "admin", "modules/admin/users.php", "Create user", "Admin logged in", "Add user with required fields", "User saved and listed", "High", "Positive", "Admin"))
    add(("F02", "admin", "modules/admin/users.php", "Edit user role", "User exists", "Update role and save", "Role updated", "High", "Positive", "Admin"))
    add(("F02", "admin", "modules/admin/active_sessions.php", "View sessions", "Admin logged in", "Open active sessions page", "Sessions listed", "Medium", "Positive", "Admin"))
    add(("F02", "admin", "modules/admin/login_activity.php", "Delete login records", "Activity exists", "Delete records with confirmation", "Records removed", "Medium", "Regression", "Admin"))

    add(("F03", "projects", "modules/projects/create.php", "Create project", "Admin/Lead", "Create project", "Project created", "High", "Positive", "Admin/Lead"))
    add(("F03", "projects", "modules/projects/edit.php", "Edit project", "Project exists", "Update project details", "Project updated", "High", "Positive", "Admin/Lead"))
    add(("F03", "projects", "modules/projects/archive.php", "Archive project", "Project exists", "Archive action", "Project archived", "High", "Regression", "Admin/Lead"))
    add(("F03", "projects", "modules/projects/duplicate.php", "Duplicate project", "Project exists", "Duplicate action", "Project copy created", "Medium", "Positive", "Admin/Lead"))
    add(("F03", "projects", "modules/projects/delete.php", "Delete project", "Project exists", "Delete action + confirm", "Project deleted", "High", "Regression", "Admin/Lead"))

    add(("F04", "projects", "modules/projects/manage_assignments.php", "Assign member", "Project exists", "Assign team member", "Assignment created", "High", "Positive", "Admin/Lead"))
    add(("F04", "projects", "modules/projects/manage_assignments.php", "Remove member", "Member assigned", "Remove team member", "Assignment removed/soft deleted", "High", "Regression", "Admin/Lead"))
    add(("F04", "projects", "modules/projects/add_page.php", "Add page", "Project exists", "Add project page", "Page added", "High", "Positive", "Admin/Lead"))
    add(("F04", "projects", "modules/projects/bulk_assign.php", "Bulk assign pages", "Pages exist", "Bulk assign testers/qa", "Assignments applied", "High", "Positive", "Admin/Lead"))
    add(("F04", "qa", "modules/qa/page_assignment.php", "QA assignment view scope", "QA logged in", "Open page assignment", "Only authorized items visible", "Medium", "Security", "QA"))

    add(("F05", "at_tester", "modules/at_tester/project_tasks.php", "AT task list", "AT assigned", "Open AT tasks", "Only assigned tasks visible", "High", "Positive", "AT"))
    add(("F05", "at_tester", "modules/at_tester/project_tasks.php", "AT task update", "AT task exists", "Update task status", "Status saved", "High", "Positive", "AT"))
    add(("F05", "ft_tester", "modules/ft_tester/project_tasks.php", "FT task list", "FT assigned", "Open FT tasks", "Only assigned tasks visible", "High", "Positive", "FT"))
    add(("F05", "ft_tester", "modules/ft_tester/project_tasks.php", "FT task update", "FT task exists", "Update task status", "Status saved", "High", "Positive", "FT"))
    add(("F05", "qa", "modules/qa/qa_tasks.php", "QA review", "QA assigned", "Update QA status", "QA status saved", "High", "Positive", "QA"))

    add(("F06", "projects", "modules/projects/issues_all.php", "Create issue", "Project/page exists", "Create issue", "Issue saved", "High", "Positive", "Assigned"))
    add(("F06", "projects", "modules/projects/issues_page_detail.php", "Page-level issue", "Page exists", "Create issue from page", "Issue linked to page", "High", "Positive", "Assigned"))
    add(("F06", "projects", "modules/projects/issues_common.php", "Common issues list", "Issues exist", "Open common issues", "Issues listed", "Medium", "Positive", "Assigned"))
    add(("F06", "admin", "modules/admin/issue_config.php", "Issue preset config", "Admin", "Create/update preset", "Config saved", "Medium", "Positive", "Admin"))
    add(("F06", "projects", "modules/projects/export_issues.php", "Issue export", "Issues exist", "Export issues", "Download generated", "Medium", "Positive", "Assigned"))

    add(("F07", "core", "modules/my_daily_status.php", "Update status", "User logged in", "Save daily status", "Status updated", "High", "Positive", "All"))
    add(("F07", "core", "modules/my_daily_status.php", "Log hours", "Project assigned", "Add time log entry", "Hours saved", "High", "Positive", "All"))
    add(("F07", "core", "modules/calendar.php", "Calendar render", "User logged in", "Open calendar", "Entries displayed correctly", "Medium", "Positive", "All"))
    add(("F07", "admin", "modules/admin/calendar.php", "Admin edit user status", "Admin", "Open modal > edit > save", "Changes saved; readonly/edit behavior correct", "High", "Regression", "Admin"))
    add(("F07", "admin", "modules/admin/production_logs.php", "Production logs filter", "Logs exist", "Filter by user/date", "Correct rows shown", "Medium", "Positive", "Admin"))

    add(("F08", "chat", "modules/chat/project_chat.php", "Send chat message", "Project access", "Send message", "Message appears", "High", "Positive", "Assigned"))
    add(("F08", "chat", "modules/chat/project_chat.php", "Upload chat attachment", "Upload permission", "Upload file in chat", "File linked and opens as expected", "High", "Positive", "Assigned"))
    add(("F08", "chat", "modules/chat/page_chat.php", "Page chat thread", "Page access", "Send page chat message", "Message appears in page thread", "Medium", "Positive", "Assigned"))
    add(("F08", "core", "modules/notifications.php", "Notification listing", "Notifications exist", "Open notifications", "Items listed and links valid", "Medium", "Positive", "All"))

    add(("F09", "reports", "modules/reports/dashboard.php", "Reports dashboard", "Reports access", "Open dashboard", "Widgets load", "Medium", "Positive", "Reports/Admin"))
    add(("F09", "reports", "modules/reports/export.php", "Export report", "Data exists", "Apply filter and export", "Export file correct", "Medium", "Positive", "Reports/Admin"))
    add(("F09", "admin", "modules/admin/reports.php", "Admin reports", "Admin", "Open admin reports", "Page loads with data", "Low", "Regression", "Admin"))

    add(("F10", "admin", "modules/admin/clients.php", "Client CRUD", "Admin", "Create/update/delete client", "Client master maintained", "Medium", "Positive", "Admin"))
    add(("F10", "admin", "modules/admin/environments.php", "Environment CRUD", "Admin", "Create/update environment", "Environment master maintained", "Medium", "Positive", "Admin"))
    add(("F10", "admin", "modules/admin/phase_master.php", "Phase master", "Admin", "Create/update phase", "Phase master maintained", "Low", "Positive", "Admin"))
    add(("F10", "admin", "modules/admin/manage_statuses.php", "Availability status master", "Admin", "Update statuses", "Status options updated", "Low", "Regression", "Admin"))

    add(("F11", "projects", "modules/projects/assets_list.php", "Add asset", "Project access", "Upload project asset", "Asset listed", "High", "Positive", "Assigned"))
    add(("F11", "projects", "modules/projects/handle_asset.php", "Delete asset", "Asset exists", "Delete with confirmation", "Asset removed and redirected correctly", "High", "Regression", "Assigned"))
    add(("F11", "admin", "modules/admin/uploads_manager.php", "Uploads manager list", "Admin", "Open uploads manager", "Uploads visible with filters", "High", "Positive", "Admin"))
    add(("F11", "admin", "modules/admin/uploads_manager.php", "Uploads scoped delete", "Records exist", "Delete by user/project scope", "Only selected scope deleted", "High", "Positive", "Admin"))

    add(("F12", "core", "modules/profile.php", "Profile assigned projects", "User has assignments", "Open profile?id={user}", "Project count/list accurate", "High", "Regression", "All"))
    add(("F12", "core", "modules/profile.php", "Profile assigned tasks", "User has tasks", "Verify Assigned Tasks section", "Task count/list accurate", "High", "Regression", "All"))
    add(("F12", "core", "modules/notifications.php", "Notification read flow", "Unread exists", "Open notification link", "Item marked read", "Medium", "Positive", "All"))

    add(("F13", "generic_tasks", "modules/generic_tasks/manage_categories.php", "Create category", "Admin", "Create category", "Category saved", "Low", "Positive", "Admin"))
    add(("F13", "generic_tasks", "modules/generic_tasks/add_task.php", "Create generic task", "Category exists", "Add task", "Task saved", "Low", "Positive", "Admin"))
    add(("F13", "generic_tasks", "modules/generic_tasks/index.php", "Search/filter tasks", "Tasks exist", "Filter and search", "Expected rows shown", "Low", "Positive", "Admin"))

    add(("F14", "core", "Delete actions", "Delete confirmation", "Delete UI exists", "Trigger delete action", "Confirmation shown before delete", "High", "Regression", "All"))
    add(("F14", "core", "Role-protected pages", "RBAC enforcement", "Different roles", "Try unauthorized URLs", "Access blocked", "High", "Security", "All"))
    add(("F14", "admin", "modules/admin/backup.php", "Backup generation", "Admin", "Generate backup", "Backup completes without errors", "Medium", "Resilience", "Admin"))
    return rows


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    for col in ws.columns:
        m = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, m + 2), 65)


def main():
    pages = discover_pages()
    modules = group_modules(pages)
    flows = get_flows()
    cases = get_cases()

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("ReadMe")
    ws.append(["Field", "Value"])
    ws.append(["Project", "PMS Manual Testing Workbook"])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Coverage", "Flow-wise, Module-wise, Page-wise, Detailed test cases"])
    ws.append(["Total Modules", len(modules)])
    ws.append(["Total Pages", len(pages)])
    ws.append(["Total Cases", len(cases)])
    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("Flow_Wise")
    ws.append(["Flow ID", "Flow Name", "Objective", "Modules", "Priority"])
    for r in flows:
        ws.append(list(r))
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    ws = wb.create_sheet("Module_Wise")
    ws.append(["Module", "Page Count", "Sample Pages", "Focus Areas"])
    focus = {
        "admin": "RBAC, masters, users, sessions, reports, uploads",
        "auth": "login/logout/reset",
        "projects": "project/page/issue/asset lifecycle",
        "project_lead": "team/page assignments",
        "at_tester": "AT task execution",
        "ft_tester": "FT task execution",
        "qa": "QA tasks and status",
        "chat": "messages and attachments",
        "reports": "dashboard/export accuracy",
        "generic_tasks": "task/category CRUD",
        "core": "calendar, profile, notifications, daily status",
    }
    for mod, mp in modules.items():
        ws.append([mod, len(mp), ", ".join(Path(x).name for x in mp[:6]), focus.get(mod, "page load + CRUD + RBAC")])
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    ws = wb.create_sheet("Page_Wise")
    ws.append(["Seq", "Module", "Page Path", "Page Name", "Priority", "Primary Checks"])
    i = 1
    for mod, mp in modules.items():
        for p in mp:
            n = Path(p).name.lower()
            pr = "High" if any(k in n for k in ["delete", "manage", "login", "users", "archive", "duplicate"]) else ("Medium" if any(k in n for k in ["dashboard", "view", "issues", "tasks"]) else "Low")
            ws.append([i, mod, p, Path(p).name, pr, "Access control, validation, data load, actions, redirect"])
            i += 1
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    ws = wb.create_sheet("Detailed_Test_Cases")
    ws.append(["TC ID", "Flow ID", "Module", "Page", "Scenario", "Precondition", "Steps", "Expected", "Priority", "Type", "Role", "Status", "Actual", "Defect ID", "Tester", "Date", "Remarks"])
    for i, (fid, mod, page, scn, pre, stp, exp, pri, typ, role) in enumerate(cases, start=1):
        ws.append([f"TC-{i:03d}", fid, mod, page, scn, pre, stp, exp, pri, typ, role, "", "", "", "", "", ""])
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 48
    ws.column_dimensions["H"].width = 48
    ws.column_dimensions["M"].width = 40
    ws.column_dimensions["Q"].width = 32
    autofit(ws)

    ws = wb.create_sheet("RTM")
    ws.append(["Flow ID", "Flow Name", "TC Count", "High Priority TCs"])
    for fid, fname, *_ in flows:
        flow_cases = [x for x in cases if x[0] == fid]
        ws.append([fid, fname, len(flow_cases), len([x for x in flow_cases if x[7] == "High"])])
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"Modules={len(modules)}, Pages={len(pages)}, Cases={len(cases)}")


if __name__ == "__main__":
    main()

