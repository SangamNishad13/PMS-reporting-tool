from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from generate_manual_test_workbook import get_cases, get_flows


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "PMS_UAT_Smoke_Critical_Test_Cases.xlsx"


def style_header(ws):
    fill = PatternFill("solid", fgColor="7A1E1E")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    for col in ws.columns:
        m = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, m + 2), 65)


def main():
    flows = get_flows()
    cases = get_cases()

    # UAT smoke scope:
    # 1) All High priority cases
    # 2) Plus key Medium regression flows for end-to-end handoff
    selected = []
    for c in cases:
        fid, mod, page, scn, pre, stp, exp, pri, typ, role = c
        if pri == "High":
            selected.append(c)
            continue
        if fid in {"F03", "F06", "F07", "F11", "F12"} and typ in {"Regression", "Security"}:
            selected.append(c)

    # De-duplicate keeping order
    seen = set()
    final_cases = []
    for c in selected:
        key = (c[0], c[2], c[3], c[4])
        if key in seen:
            continue
        seen.add(key)
        final_cases.append(c)

    flow_names = {f[0]: f[1] for f in flows}
    involved_flows = sorted(set(c[0] for c in final_cases))

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("ReadMe")
    ws.append(["Field", "Value"])
    ws.append(["Suite", "PMS UAT Smoke + Critical Path"])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Selection Rule", "All High priority + critical Medium regression/security cases"])
    ws.append(["Total Selected Cases", len(final_cases)])
    ws.append(["Flows Covered", ", ".join(involved_flows)])
    ws.append(["Execution Target", "Manual testing handoff for pre-live/UAT signoff"])
    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("Smoke_Flows")
    ws.append(["Flow ID", "Flow Name", "Why Critical For UAT"])
    for fid in involved_flows:
        ws.append([fid, flow_names.get(fid, fid), "Core business path / release blocker coverage"])
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    ws = wb.create_sheet("Critical_Path_TCs")
    ws.append([
        "TC ID", "Flow ID", "Flow Name", "Module", "Page", "Scenario",
        "Precondition", "Steps", "Expected", "Priority", "Type", "Role",
        "Status", "Actual", "Defect ID", "Tester", "Date", "Remarks"
    ])
    for i, c in enumerate(final_cases, start=1):
        fid, mod, page, scn, pre, stp, exp, pri, typ, role = c
        ws.append([
            f"UAT-TC-{i:03d}",
            fid, flow_names.get(fid, fid), mod, page, scn,
            pre, stp, exp, pri, typ, role,
            "", "", "", "", "", ""
        ])
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 48
    ws.column_dimensions["I"].width = 48
    ws.column_dimensions["N"].width = 40
    ws.column_dimensions["R"].width = 32
    autofit(ws)

    ws = wb.create_sheet("Execution_Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Total Test Cases", len(final_cases)])
    ws.append(["Passed", ""])
    ws.append(["Failed", ""])
    ws.append(["Blocked", ""])
    ws.append(["Not Run", ""])
    ws.append(["Open Defects", ""])
    ws.append(["Go/No-Go Recommendation", ""])
    style_header(ws)
    autofit(ws)

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"Selected UAT cases: {len(final_cases)}")


if __name__ == "__main__":
    main()

