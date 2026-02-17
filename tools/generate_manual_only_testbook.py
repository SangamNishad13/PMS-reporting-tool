from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from generate_manual_test_workbook import get_cases, get_flows


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "PMS_Manual_Test_Scenarios_and_Cases_Only.xlsx"


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    for col in ws.columns:
        m = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, m + 2), 70)


def main():
    flows = get_flows()
    flow_map = {f[0]: f[1] for f in flows}
    cases = get_cases()

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Test_Scenarios")
    ws.append(["Scenario ID", "Flow ID", "Scenario Name", "Module", "Page", "Priority", "Type", "Role"])
    for i, c in enumerate(cases, start=1):
        fid, mod, page, scn, pre, stp, exp, pri, typ, role = c
        ws.append([f"TS-{i:03d}", fid, scn, mod, page, pri, typ, role])
    style_header(ws)
    ws.freeze_panes = "A2"
    autofit(ws)

    ws = wb.create_sheet("Test_Cases")
    ws.append([
        "TC ID", "Flow ID", "Flow Name", "Scenario",
        "Precondition", "Steps", "Expected Result",
        "Priority", "Type", "Role",
        "Status", "Actual Result", "Defect ID", "Tester", "Test Date", "Remarks"
    ])
    for i, c in enumerate(cases, start=1):
        fid, mod, page, scn, pre, stp, exp, pri, typ, role = c
        ws.append([
            f"TC-{i:03d}", fid, flow_map.get(fid, fid), scn,
            pre, stp, exp, pri, typ, role,
            "", "", "", "", "", ""
        ])
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 52
    ws.column_dimensions["G"].width = 52
    ws.column_dimensions["L"].width = 38
    ws.column_dimensions["P"].width = 30
    autofit(ws)

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"Scenarios/Cases: {len(cases)}")
    print(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()

